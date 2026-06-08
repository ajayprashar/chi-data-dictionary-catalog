#!/usr/bin/env python3
"""Import CHI steward workbook data sheets back into repo parquet files."""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


SHEET_TO_PARQUET: dict[str, str] = {
    "Catalog": "ddc-master_patient_catalog.parquet",
    "Dictionary": "ddc-master_patient_dictionary.parquet",
    "Source_Availability": "ddc-data_source_availability.parquet",
    "ADT_Mappings": "ddc-hl7_adt_catalog.parquet",
    "CCDA_Mappings": "ddc-ccda_catalog.parquet",
    "FHIR_Inventory": "ddc-fhir_inventory.parquet",
    "Business_Rules": "ddc-business_rules.parquet",
}


def sheet_to_dataframe(ws) -> pd.DataFrame:
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame()
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        data_rows.append(["" if cell is None else str(cell) for cell in row])
    return pd.DataFrame(data_rows, columns=headers)


def import_workbook(workbook_path: Path, repo_root: Path) -> None:
    wb = load_workbook(workbook_path, data_only=True)
    for sheet_name, parquet_name in SHEET_TO_PARQUET.items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Workbook is missing required sheet: {sheet_name}")
        df = sheet_to_dataframe(wb[sheet_name])
        out_path = repo_root / parquet_name
        df.to_parquet(out_path, index=False)
        print(f"Wrote {out_path} ({len(df)} rows) from sheet {sheet_name}")


def main() -> None:
    repo_root = Path(__file__).resolve().parent.parent
    default_workbook = repo_root / "workbooks" / "chi-steward-workbook.xlsx"

    parser = argparse.ArgumentParser(description="Import steward workbook sheets into parquet files.")
    parser.add_argument(
        "workbook",
        nargs="?",
        default=str(default_workbook),
        help=f"Path to steward workbook (default: {default_workbook})",
    )
    parser.add_argument(
        "-d",
        "--repo-root",
        default=str(repo_root),
        help="Project root containing ddc-*.parquet outputs",
    )
    args = parser.parse_args()

    import_workbook(Path(args.workbook), Path(args.repo_root))


if __name__ == "__main__":
    main()
