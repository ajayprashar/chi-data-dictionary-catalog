#!/usr/bin/env python3
"""
Generate the CHI steward workbook from repo parquet files.

Each governed data sheet is a named Excel Table (chi_* prefix) aligned to
ddc-* parquet artifacts. Concept_Explorer uses structured references.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).resolve().parent))
from excel_workbook_common import (
    MODEL_FILL,
    add_excel_table,
    autosize_columns,
    style_header_row,
    table_lookup_formula,
    write_readme_sheet,
)


OUT_NAME = "chi-steward-workbook.xlsx"

# (parquet file, sheet tab name, Excel Table name)
DATA_TABLES: list[tuple[str, str, str]] = [
    ("ddc-master_patient_catalog.parquet", "Catalog", "chi_catalog"),
    ("ddc-master_patient_dictionary.parquet", "Dictionary", "chi_dictionary"),
    ("ddc-data_source_availability.parquet", "Source_Availability", "chi_source_availability"),
    ("ddc-hl7_adt_catalog.parquet", "ADT_Mappings", "chi_adt_mappings"),
    ("ddc-ccda_catalog.parquet", "CCDA_Mappings", "chi_ccda_mappings"),
    ("ddc-fhir_inventory.parquet", "FHIR_Inventory", "chi_fhir_inventory"),
    ("ddc-business_rules.parquet", "Business_Rules", "chi_business_rules"),
    ("ddc-value_set_binding.parquet", "Value_Set_Bindings", "chi_value_set_binding"),
    ("ddc-value_set_member.parquet", "Value_Set_Members", "chi_value_set_member"),
    ("ddc-source_value_crosswalk.parquet", "Source_Value_Crosswalk", "chi_source_value_crosswalk"),
]

SHEET_TO_TABLE = {sheet: table for _, sheet, table in DATA_TABLES}
SHEET_TO_TABLE["Source_Registry"] = "chi_source_registry"
SHEET_TO_TABLE["Steward_Queue"] = "chi_steward_queue"
SHEET_TO_TABLE["Lookup_Lists"] = "chi_lookup_lists"

MODEL_ROWS = [
    (
        "rel_id",
        "parent_table",
        "parent_key",
        "child_table",
        "child_key",
        "cardinality",
        "description",
    ),
    (
        "catalog_dictionary",
        "chi_catalog",
        "semantic_id",
        "chi_dictionary",
        "semantic_id",
        "1:1",
        "Each governed concept has one implementation definition row.",
    ),
    (
        "catalog_source_availability",
        "chi_catalog",
        "semantic_id",
        "chi_source_availability",
        "semantic_id",
        "1:N",
        "Links each concept to the data sources that can provide it.",
    ),
    (
        "catalog_adt",
        "chi_catalog",
        "semantic_id",
        "chi_adt_mappings",
        "semantic_id",
        "1:N",
        "HL7 ADT segment/field placements for the same concept.",
    ),
    (
        "catalog_ccda",
        "chi_catalog",
        "semantic_id",
        "chi_ccda_mappings",
        "semantic_id",
        "1:N",
        "C-CDA/CCD XML paths for the same concept.",
    ),
    (
        "catalog_fhir_inventory",
        "chi_catalog",
        "semantic_id",
        "chi_fhir_inventory",
        "semantic_id",
        "1:N",
        "FHIR standards inventory rows tied to the concept.",
    ),
    (
        "catalog_business_rules",
        "chi_catalog",
        "semantic_id",
        "chi_business_rules",
        "semantic_id",
        "1:N",
        "Organization-specific rules for the concept.",
    ),
    (
        "catalog_value_set_binding",
        "chi_catalog",
        "semantic_id",
        "chi_value_set_binding",
        "semantic_id",
        "1:N",
        "Which terminology / value set applies to the concept.",
    ),
    (
        "catalog_value_set_member",
        "chi_catalog",
        "semantic_id",
        "chi_value_set_member",
        "semantic_id",
        "1:N",
        "Governed codes for the concept (CHI subset).",
    ),
    (
        "catalog_source_crosswalk",
        "chi_catalog",
        "semantic_id",
        "chi_source_value_crosswalk",
        "semantic_id",
        "1:N",
        "Source-local values mapped to governed standard codes.",
    ),
    (
        "catalog_steward_queue",
        "chi_catalog",
        "semantic_id",
        "chi_steward_queue",
        "semantic_id",
        "1:1",
        "Optional steward workflow overlay keyed by semantic_id.",
    ),
]

CONCEPT_EXPLORER_DEFAULT_SEMANTIC_ID = "Patient.race"

DEMOGRAPHICS_PILOT: dict[str, dict[str, str]] = {
    "Patient.race": {
        "curation_status": "Approved",
        "steward_assigned_to": "Ajay Prashar",
        "steward_action_notes": "Health equity dashboard; ASCMI consent; crosswalk validation",
        "priority": "POC",
        "queue_type": "demographics_pilot",
        "next_action": "Done",
    },
    "Patient.ethnicity": {
        "curation_status": "Approved",
        "steward_assigned_to": "Ajay Prashar",
        "steward_action_notes": "Health equity dashboard; ASCMI consent",
        "priority": "POC",
        "queue_type": "demographics_pilot",
        "next_action": "Done",
    },
    "Patient.language": {
        "curation_status": "Approved",
        "steward_assigned_to": "Ajay Prashar",
        "steward_action_notes": "Health equity dashboard; outreach language targeting",
        "priority": "POC",
        "queue_type": "demographics_pilot",
        "next_action": "Done",
    },
    "Patient.gender_id": {
        "curation_status": "Approved",
        "steward_assigned_to": "Ajay Prashar",
        "steward_action_notes": "ASCMI consent; equity reporting",
        "priority": "POC",
        "queue_type": "demographics_pilot",
        "next_action": "Done",
    },
    "Patient.birth_sex": {
        "curation_status": "Approved",
        "steward_assigned_to": "Ajay Prashar",
        "steward_action_notes": "UMPI matching; clinical use; not interchangeable with gender identity",
        "priority": "POC",
        "queue_type": "demographics_pilot",
        "next_action": "Done",
    },
}

STEWARD_QUEUE_HEADERS = [
    "semantic_id",
    "curation_status",
    "steward_assigned_to",
    "steward_action_notes",
    "reviewer_notes",
    "priority",
    "queue_type",
    "next_action",
]

# (Excel Table name, column name, friendly label)
EXPLORER_FIELDS = [
    ("chi_catalog", "uscdi_element", "USCDI element"),
    ("chi_catalog", "uscdi_description", "USCDI description"),
    ("chi_catalog", "classification", "Classification"),
    ("chi_catalog", "data_steward", "Data steward"),
    ("chi_catalog", "approval_status", "Approval status"),
    ("chi_dictionary", "fhir_r4_path", "FHIR R4 path"),
    ("chi_dictionary", "fhir_profile", "US Core profile"),
    ("chi_dictionary", "data_quality_notes", "Terminology / standards notes"),
    ("chi_dictionary", "chi_survivorship_logic", "CHI survivorship logic"),
    ("chi_dictionary", "data_source_rank_reference", "Data source rank"),
    ("chi_adt_mappings", "field_id", "HL7 ADT field (first match)"),
    ("chi_adt_mappings", "segment_id", "HL7 ADT segment (first match)"),
    ("chi_ccda_mappings", "xml_path", "C-CDA XML path (first match)"),
    ("chi_value_set_member", "code", "Governed code (first member)"),
    ("chi_value_set_member", "display", "Governed code display (first member)"),
    ("chi_source_value_crosswalk", "source_code", "Source crosswalk code (first match)"),
    ("chi_source_availability", "source_id", "Primary source (first link)"),
    ("chi_source_availability", "availability", "Source availability"),
]

# (excel_table_name, sheet_tab, parquet_or_role, role)
TABLE_README_ROWS = [
    ("(support)", "Concept_Explorer", "(lookup view; not imported)", "Start here"),
    ("chi_catalog", "Catalog", "ddc-master_patient_catalog.parquet", "Start here"),
    ("chi_dictionary", "Dictionary", "ddc-master_patient_dictionary.parquet", "Start here"),
    ("chi_source_availability", "Source_Availability", "ddc-data_source_availability.parquet", "Source linking"),
    ("chi_source_registry", "Source_Registry", "(steward metadata)", "Source linking"),
    ("chi_steward_queue", "Steward_Queue", "(Excel workflow overlay)", "Workflow"),
    ("chi_adt_mappings", "ADT_Mappings", "ddc-hl7_adt_catalog.parquet", "Mappings"),
    ("chi_ccda_mappings", "CCDA_Mappings", "ddc-ccda_catalog.parquet", "Mappings"),
    ("chi_fhir_inventory", "FHIR_Inventory", "ddc-fhir_inventory.parquet", "Mappings"),
    ("chi_business_rules", "Business_Rules", "ddc-business_rules.parquet", "Mappings"),
    ("chi_value_set_binding", "Value_Set_Bindings", "ddc-value_set_binding.parquet", "Terminology"),
    ("chi_value_set_member", "Value_Set_Members", "ddc-value_set_member.parquet", "Terminology"),
    ("chi_source_value_crosswalk", "Source_Value_Crosswalk", "ddc-source_value_crosswalk.parquet", "Crosswalk"),
    ("chi_lookup_lists", "Lookup_Lists", "(reference lists)", "Reference"),
    ("chi_table_index", "Table_Index", "(navigation map; not imported)", "Reference"),
]


def load_parquet_frames(repo_root: Path) -> dict[str, pd.DataFrame]:
    frames: dict[str, pd.DataFrame] = {}
    for parquet_name, _, _ in DATA_TABLES:
        path = repo_root / parquet_name
        if not path.exists():
            raise FileNotFoundError(f"Missing required parquet file: {path}")
        frames[parquet_name] = pd.read_parquet(path).fillna("")
    return frames


def write_dataframe_sheet(ws: Worksheet, df: pd.DataFrame, table_name: str) -> list[str]:
    headers = [str(col) for col in df.columns]
    ws.append(headers)
    for row in df.itertuples(index=False, name=None):
        ws.append(["" if value is None else str(value) for value in row])
    style_header_row(ws)
    add_excel_table(ws, table_name)
    autosize_columns(ws)
    return headers


def add_model_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("_Model")
    for row in MODEL_ROWS:
        ws.append(row)
    style_header_row(ws)
    for cell in ws[1]:
        cell.fill = MODEL_FILL
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    autosize_columns(ws)
    ws.sheet_state = "hidden"


def add_table_index_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Table_Index")
    ws.append(("excel_table_name", "sheet_tab", "parquet_or_role", "role"))
    for row in TABLE_README_ROWS:
        ws.append(row)
    style_header_row(ws)
    add_excel_table(ws, "chi_table_index")
    autosize_columns(ws)


def add_lookup_lists_sheet(wb: Workbook, frames: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("Lookup_Lists")
    catalog = frames["ddc-master_patient_catalog.parquet"]
    availability = frames["ddc-data_source_availability.parquet"]

    semantic_ids = sorted({str(v).strip() for v in catalog["semantic_id"] if str(v).strip()})
    source_ids = sorted({str(v).strip() for v in availability["source_id"] if str(v).strip()})

    headers = [
        "semantic_id",
        "source_id",
        "mapping_status",
        "approval_status",
        "availability",
        "curation_status",
    ]
    ws.append(headers)
    style_header_row(ws)

    mapping_status = ["mapped", "needs_new_semantic", "out_of_scope"]
    approval_status = ["Pending", "Ready", "Approved", "Needs Changes"]
    availability_vals = ["full", "partial", "none", "unknown"]
    curation_status = ["Needs Action", "In Progress", "Ready for Review", "Complete"]

    max_len = max(
        len(semantic_ids),
        len(source_ids),
        len(mapping_status),
        len(approval_status),
        len(availability_vals),
        len(curation_status),
    )
    for idx in range(max_len):
        ws.append([
            semantic_ids[idx] if idx < len(semantic_ids) else "",
            source_ids[idx] if idx < len(source_ids) else "",
            mapping_status[idx] if idx < len(mapping_status) else "",
            approval_status[idx] if idx < len(approval_status) else "",
            availability_vals[idx] if idx < len(availability_vals) else "",
            curation_status[idx] if idx < len(curation_status) else "",
        ])

    add_excel_table(ws, "chi_lookup_lists")
    autosize_columns(ws)


def add_source_registry_sheet(wb: Workbook, frames: dict[str, pd.DataFrame]) -> None:
    availability = frames["ddc-data_source_availability.parquet"]
    source_ids = sorted({str(v).strip() for v in availability["source_id"] if str(v).strip()})
    ws = wb.create_sheet("Source_Registry")
    headers = ["source_id", "source_name", "organization", "delivery_format", "notes"]
    ws.append(headers)
    for source_id in source_ids:
        ws.append([source_id, "", "", "", "Populate source metadata for steward reference."])
    style_header_row(ws)
    add_excel_table(ws, "chi_source_registry")
    autosize_columns(ws)


def steward_queue_row(semantic_id: str) -> list[str]:
    pilot = DEMOGRAPHICS_PILOT.get(semantic_id, {})
    return [
        semantic_id,
        pilot.get("curation_status", ""),
        pilot.get("steward_assigned_to", ""),
        pilot.get("steward_action_notes", ""),
        pilot.get("reviewer_notes", ""),
        pilot.get("priority", ""),
        pilot.get("queue_type", ""),
        pilot.get("next_action", ""),
    ]


def add_steward_queue_sheet(wb: Workbook, frames: dict[str, pd.DataFrame]) -> None:
    catalog = frames["ddc-master_patient_catalog.parquet"]
    catalog_ids = [str(v).strip() for v in catalog["semantic_id"] if str(v).strip()]
    pilot_ids = [sid for sid in DEMOGRAPHICS_PILOT if sid in catalog_ids]
    other_ids = sorted(sid for sid in catalog_ids if sid not in DEMOGRAPHICS_PILOT)

    ws = wb.create_sheet("Steward_Queue")
    ws.append(STEWARD_QUEUE_HEADERS)
    for semantic_id in pilot_ids + other_ids:
        ws.append(steward_queue_row(semantic_id))
    style_header_row(ws)
    add_excel_table(ws, "chi_steward_queue")
    autosize_columns(ws)


def add_concept_explorer_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Concept_Explorer")
    ws["A1"] = "Concept Explorer"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = (
        "Change semantic_id in B3 to preview catalog, dictionary, ADT/CCDA context, and sources. "
        "ADT/CCDA show first matching row; see ADT_Mappings and CCDA_Mappings for full lists."
    )
    ws["A3"] = "semantic_id"
    ws["B3"] = CONCEPT_EXPLORER_DEFAULT_SEMANTIC_ID
    ws["A4"] = "Field"
    ws["B4"] = "Value"
    ws["A4"].font = Font(bold=True)
    ws["B4"].font = Font(bold=True)

    start_row = 5
    for offset, (table_name, column_name, label) in enumerate(EXPLORER_FIELDS):
        row = start_row + offset
        ws.cell(row=row, column=1, value=label)
        ws.cell(
            row=row,
            column=2,
            value=table_lookup_formula(table_name, "semantic_id", column_name),
        )

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 72
    for row in ws.iter_rows(min_row=5, max_row=start_row + len(EXPLORER_FIELDS) - 1, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_readme_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    pilot_ids = ", ".join(DEMOGRAPHICS_PILOT)
    write_readme_sheet(
        ws,
        "CHI Steward Workbook",
        [
            (
                "Purpose",
                "Governed data catalog + dictionary over ddc-* parquet, curated to healthcare standards, "
                "with HL7 ADT and C-CDA context sheets. See docs/product-vision.md.",
            ),
            (
                "Table naming",
                "Pattern: chi_{artifact} in snake_case. Examples: chi_catalog, chi_dictionary, chi_source_availability.",
            ),
            (
                "Table_Index",
                "Maps every sheet to role (Start here, Source linking, Mappings, Workflow, Reference), "
                "Excel Table name, and parquet artifact.",
            ),
            ("Join key", "semantic_id links chi_catalog to chi_dictionary, chi_source_availability, and mapping tables."),
            (
                "Concept_Explorer",
                "Change semantic_id in B3 to preview catalog, dictionary, and source fields on one row.",
            ),
            (
                "Steward_Queue",
                "Demographics pilot concepts are listed first with suggested next actions. "
                "Track curation_status and steward notes here (not imported to parquet).",
            ),
            (
                "Round trip",
                "python scripts/generate_steward_workbook.py exports parquet to Excel. "
                "python scripts/import_steward_workbook_to_parquet.py writes edits back to parquet.",
            ),
            ("Do not import", "README, _Model, Table_Index, Lookup_Lists, Concept_Explorer are support sheets."),
        ],
        start_here_steps=[
            "Open Concept_Explorer - B3 defaults to Patient.race; try the other demographics pilot IDs next.",
            f"Demographics pilot: {pilot_ids}.",
            "Edit Catalog and Dictionary for each concept; use Lookup_Lists for allowed status values.",
            "Check Source_Availability for source_id links; use Steward_Queue for workflow notes.",
            "When ready: python scripts/import_steward_workbook_to_parquet.py",
        ],
    )


def main() -> None:
    repo_root = Path(__file__).resolve().parent.parent
    out_dir = repo_root / "workbooks"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / OUT_NAME

    frames = load_parquet_frames(repo_root)

    wb = Workbook()
    add_readme_sheet(wb)
    add_model_sheet(wb)
    add_table_index_sheet(wb)
    add_lookup_lists_sheet(wb, frames)

    for parquet_name, sheet_name, table_name in DATA_TABLES:
        ws = wb.create_sheet(sheet_name)
        write_dataframe_sheet(ws, frames[parquet_name], table_name)

    add_source_registry_sheet(wb, frames)
    add_steward_queue_sheet(wb, frames)
    add_concept_explorer_sheet(wb)

    wb.save(out_path)
    print(f"Wrote steward workbook: {out_path}")
    for parquet_name, sheet_name, table_name in DATA_TABLES:
        rows = len(frames[parquet_name])
        print(f"  - {sheet_name} -> {table_name}: {rows} rows from {parquet_name}")


if __name__ == "__main__":
    main()
