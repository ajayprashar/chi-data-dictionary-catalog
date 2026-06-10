#!/usr/bin/env python3
"""
Generate the CHI partner intake workbook.

Each data sheet is a named Excel Table (chi_intake_* prefix). Dropdown
validations are omitted because openpyxl validations trigger Excel repair
prompts on open.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(REPO_ROOT))
from data.partner_crosswalk_template import CROSSWALK_COLUMNS, build_template_rows  # noqa: E402
from excel_workbook_common import (
    add_excel_table,
    autosize_columns,
    style_header_row,
    write_readme_sheet,
)


OUT_NAME = "chi-partner-intake-workbook.xlsx"

PARTNER_SHEETS = {
    "Source_Summary",
    "Field_Inventory",
    "Code_Values",
    "Crosswalk_Template",
    "Keys_and_Relationships",
    "Open_Questions",
}

# (excel_table_name, sheet_tab, filled_by, role)
TABLE_INDEX_ROWS = [
    ("(support)", "Intake_Guide", "partner", "Start here"),
    ("chi_intake_source_summary", "Source_Summary", "Partner", "Intake flow"),
    ("chi_intake_field_inventory", "Field_Inventory", "Partner", "Intake flow"),
    ("chi_intake_code_values", "Code_Values", "Partner", "Intake flow"),
    ("chi_intake_crosswalk_template", "Crosswalk_Template", "Partner", "Intake flow"),
    ("chi_intake_keys_relationships", "Keys_and_Relationships", "Partner", "Intake flow"),
    ("chi_intake_open_questions", "Open_Questions", "Partner", "Intake flow"),
    ("chi_intake_curation_bridge", "CHI_Curation_Bridge", "CHI", "CHI bridge"),
    ("chi_intake_governance_load_plan", "Governance_Load_Plan", "CHI", "CHI bridge"),
    ("chi_intake_lookup_lists", "Lookup_Lists", "reference", "Reference"),
    ("chi_intake_table_index", "Table_Index", "reference", "Reference"),
]

FIELD_INVENTORY_GUIDE_ROWS = [
    ("source_field_name", "Exact column or element name in the partner export."),
    ("source_table_or_file", "File, table, or message segment that contains the field."),
    ("business_label", "Plain-language label your staff uses for the field."),
    ("description", "What the field means and when it is populated."),
    ("data_type", "string, integer, date, datetime, boolean, code, etc."),
    ("required_optional", "Required or Optional in the source system."),
    ("example_value", "One real or realistic sample value."),
    ("allowed_values_or_code_set", "Code list name or short value examples; leave blank if free text."),
    ("contains_identifier", "true if the field can identify a person (copy from Lookup_Lists)."),
    ("contains_sensitive_data", "true if HIPAA-sensitive or similarly protected."),
    ("updated_when", "When the value is set or changed in the source."),
    ("notes", "Join keys, history behavior, or mapping hints for CHI."),
]

def _crosswalk_template_intake_rows() -> list[list[str]]:
    """Partner-facing rows: same columns as steward Source_Value_Crosswalk."""
    return [[row[col] for col in CROSSWALK_COLUMNS] for row in build_template_rows()]


INTAKE_TABLES: list[tuple[str, str, list[str], list[list[str]]]] = [
    (
        "Source_Summary",
        "chi_intake_source_summary",
        [
            "source_id",
            "source_name",
            "organization_name",
            "business_owner",
            "technical_contact",
            "delivery_format",
            "refresh_cadence",
            "primary_record_grain",
            "primary_identifiers",
            "notes",
        ],
        [[
            "local_jail",
            "Jail Scheduling Export",
            "Local Jail",
            "Jail Operations",
            "integration.contact@example.org",
            "CSV export",
            "Daily",
            "booking",
            "BookingNumber; PersonID",
            "Replace this example row with real source details.",
        ]],
    ),
    (
        "Field_Inventory",
        "chi_intake_field_inventory",
        [
            "source_field_name",
            "source_table_or_file",
            "business_label",
            "description",
            "data_type",
            "required_optional",
            "example_value",
            "allowed_values_or_code_set",
            "contains_identifier",
            "contains_sensitive_data",
            "updated_when",
            "notes",
        ],
        [
            [
                "BookingNumber",
                "bookings.csv",
                "Booking Number",
                "Unique booking identifier for the incarceration event",
                "string",
                "Required",
                "BK-2026-000145",
                "",
                "true",
                "false",
                "At booking creation",
                "Primary booking-level key.",
            ],
            [
                "BookingDate",
                "bookings.csv",
                "Booking Date",
                "Date and time the person was booked into custody",
                "datetime",
                "Required",
                "2026-03-24 21:17:00",
                "",
                "false",
                "false",
                "At booking creation",
                "Useful for event or encounter mapping.",
            ],
            [
                "RaceCode",
                "person.csv",
                "Race Code",
                "Race category code for the person at booking",
                "code",
                "Optional",
                "A",
                "Local race list; map to CHI standards in Crosswalk_Template",
                "false",
                "true",
                "At intake or booking",
                "Example demographics field. Add ethnicity, language, and sex fields the same way.",
            ],
            [
                "GenderIdentity",
                "person.csv",
                "Gender Identity",
                "Self-reported gender identity (not sex at birth / SexID)",
                "code",
                "Optional",
                "FEMALE",
                "See Crosswalk_Template — maps to Patient.gender_id, not birth sex",
                "false",
                "true",
                "At intake when collected",
                "Distinct from SexID or administrative sex fields.",
            ],
        ],
    ),
    (
        "Code_Values",
        "chi_intake_code_values",
        [
            "field_name",
            "code_system_name",
            "local_code",
            "meaning",
            "active_inactive",
            "notes",
        ],
        [[
            "FacilityCode",
            "Jail Facility Codes",
            "SANTA_RITA",
            "Santa Rita Jail",
            "active",
            "Example code row.",
        ]],
    ),
    (
        "Crosswalk_Template",
        "chi_intake_crosswalk_template",
        list(CROSSWALK_COLUMNS),
        _crosswalk_template_intake_rows(),
    ),
    (
        "Keys_and_Relationships",
        "chi_intake_keys_relationships",
        [
            "record_type",
            "primary_key",
            "foreign_key",
            "joins_to_record_type",
            "one_to_many_or_one_to_one",
            "history_or_audit_field",
            "notes",
        ],
        [[
            "Booking",
            "BookingNumber",
            "PersonID",
            "Person",
            "many_to_one",
            "LastUpdatedDateTime",
            "A person may have multiple bookings over time.",
        ]],
    ),
    (
        "Open_Questions",
        "chi_intake_open_questions",
        [
            "topic",
            "question",
            "owner",
            "status",
            "resolution_notes",
        ],
        [[
            "Identifiers",
            "Is PersonID stable across re-bookings and historical extracts?",
            "Local Jail + CHI",
            "open",
            "",
        ]],
    ),
    (
        "CHI_Curation_Bridge",
        "chi_intake_curation_bridge",
        [
            "source_field_name",
            "proposed_semantic_id",
            "proposed_uscdi_data_class",
            "proposed_uscdi_data_element",
            "proposed_uscdi_element",
            "proposed_fhir_r4_path",
            "proposed_domain",
            "proposed_ruleset_category",
            "proposed_target_artifact",
            "curation_disposition",
            "needs_business_rule_review",
            "chi_notes",
        ],
        [[
            "BookingDate",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "needs_review",
            "true",
            "CHI completes this after intake review.",
        ]],
    ),
    (
        "Governance_Load_Plan",
        "chi_intake_governance_load_plan",
        [
            "source_field_name",
            "approved_semantic_id",
            "target_governance_table",
            "target_column_or_purpose",
            "load_action",
            "dependencies_or_prereqs",
            "notes",
        ],
        [[
            "BookingDate",
            "",
            "",
            "",
            "create_or_update",
            "Needs approved semantic mapping first.",
            "Internal CHI planning sheet for governance load steps.",
        ]],
    ),
]

LOOKUP_ROWS = [
    ("field", "allowed_values"),
    ("contains_identifier / contains_sensitive_data / needs_business_rule_review", "true | false"),
    ("status", "open | in_progress | resolved"),
    ("curation_disposition", "map_existing | needs_new_semantic | out_of_scope | needs_review"),
    (
        "proposed_domain",
        "Domain 1: Master Demographics | Domain 2: Master Patient Attributes | Domain 3: Clinical Governance",
    ),
    (
        "target_governance_table",
        "ddc-master_patient_catalog | ddc-master_patient_dictionary | ddc-hl7_adt_catalog | "
        "ddc-ccda_catalog | ddc-data_source_availability | ddc-fhir_inventory | ddc-business_rules",
    ),
    ("load_action", "create_or_update"),
    ("mapping_type", "exact | rollup | exclude"),
    ("approval_status", "draft | Approved"),
    (
        "target_governance_table (crosswalk)",
        "ddc-source_value_crosswalk (Source_Value_Crosswalk sheet in steward workbook)",
    ),
]


def write_table_sheet(ws: Worksheet, headers: list[str], rows: list[list[str]], table_name: str) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header_row(ws)
    add_excel_table(ws, table_name)
    autosize_columns(ws)


def add_lookup_lists_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Lookup_Lists")
    ws.append(["field", "allowed_values"])
    for row in LOOKUP_ROWS[1:]:
        ws.append(list(row))
    style_header_row(ws)
    add_excel_table(ws, "chi_intake_lookup_lists")
    autosize_columns(ws)


def add_table_index_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Table_Index")
    ws.append(["excel_table_name", "sheet_tab", "filled_by", "role"])
    for row in TABLE_INDEX_ROWS:
        ws.append(list(row))
    style_header_row(ws)
    add_excel_table(ws, "chi_intake_table_index")
    autosize_columns(ws)


def add_intake_guide_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Intake_Guide")
    ws["A1"] = "Partner Intake Guide"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = (
        "Work through the Intake flow sheets in order. Replace example rows with real source details. "
        "CHI completes CHI bridge sheets after intake review."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    flow_rows = [
        ("1", "Source_Summary", "One row per source: contacts, format, refresh cadence, identifiers."),
        ("2", "Field_Inventory", "One row per field you may share. See column guide below."),
        ("3", "Code_Values", "Local code lists referenced by fields in Field_Inventory."),
        (
            "4",
            "Crosswalk_Template",
            "How your local codes map to CHI standards (example: GenderIdentity → Patient.gender_id). "
            "Replace example rows; keep source_id as partner_intake until CHI assigns yours.",
        ),
        ("5", "Keys_and_Relationships", "How records join (booking to person, encounter to patient, etc.)."),
        ("6", "Open_Questions", "Blockers and clarifications for CHI follow-up."),
    ]
    ws["A4"] = "Intake flow"
    ws["A4"].font = Font(bold=True)
    ws["A5"] = "Step"
    ws["B5"] = "Sheet"
    ws["C5"] = "What to enter"
    for cell in (ws["A5"], ws["B5"], ws["C5"]):
        cell.font = Font(bold=True)
    for offset, (step, sheet, detail) in enumerate(flow_rows):
        row = 6 + offset
        ws.cell(row=row, column=1, value=step)
        ws.cell(row=row, column=2, value=sheet)
        ws.cell(row=row, column=3, value=detail)

    guide_start = 12
    ws.cell(row=guide_start, column=1, value="Field_Inventory column guide").font = Font(bold=True)
    ws.cell(row=guide_start + 1, column=1, value="Column").font = Font(bold=True)
    ws.cell(row=guide_start + 1, column=2, value="What to enter").font = Font(bold=True)
    for offset, (column_name, description) in enumerate(FIELD_INVENTORY_GUIDE_ROWS):
        row = guide_start + 2 + offset
        ws.cell(row=row, column=1, value=column_name)
        ws.cell(row=row, column=2, value=description)

    checklist_start = guide_start + 2 + len(FIELD_INVENTORY_GUIDE_ROWS) + 1
    ws.cell(row=checklist_start, column=1, value="Before you submit").font = Font(bold=True)
    checklist_items = [
        "Source_Summary has a real source_id, business owner, and technical contact.",
        "Field_Inventory lists every identifier, demographics, and clinical field you plan to share.",
        "Code_Values documents any local code lists referenced in Field_Inventory.",
        "Crosswalk_Template maps local codes to CHI semantic_id and standard target codes (see docs/partner-crosswalk-template.md).",
        "Keys_and_Relationships explains how person, encounter, and event records join.",
        "Open_Questions has no unresolved blockers marked open without an owner.",
        "Example rows (jail booking sample) are replaced or clearly marked as examples.",
    ]
    for offset, item in enumerate(checklist_items):
        ws.cell(row=checklist_start + 1 + offset, column=1, value=f"  - {item}")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 72
    for row in ws.iter_rows(min_row=2, max_row=checklist_start + len(checklist_items), min_col=1, max_col=3):
        for cell in row:
            if cell.value:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_readme_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    partner_sheets = ", ".join(sorted(PARTNER_SHEETS))
    write_readme_sheet(
        ws,
        "CHI Partner Intake Workbook",
        [
            (
                "Purpose",
                "Gather source-system information from a partner, then bridge into CHI governance.",
            ),
            (
                "Table naming",
                "Pattern: chi_intake_{artifact}. See Table_Index for sheet roles and who fills each table.",
            ),
            (
                "Intake_Guide",
                "Partner entry sheet: intake flow steps, Field_Inventory column guide, and submission checklist.",
            ),
            (
                "Who fills what",
                f"Partner: {partner_sheets}. CHI: CHI_Curation_Bridge, Governance_Load_Plan.",
            ),
            (
                "Allowed values",
                "Copy from Lookup_Lists instead of dropdowns (avoids Excel repair prompts).",
            ),
            (
                "Boundary",
                "Partners do not assign semantic_id, FHIR paths, USCDI mappings, or survivorship logic.",
            ),
        ],
        start_here_steps=[
            "Open Intake_Guide for the step-by-step flow and Field_Inventory column definitions.",
            "Complete Source_Summary with real source metadata (replace the jail example row).",
            "Add one Field_Inventory row per export field; use the RaceCode example as a demographics template.",
            "Document code lists in Code_Values; add crosswalk rows in Crosswalk_Template where you know local → standard mappings.",
            "Record joins in Keys_and_Relationships.",
            "Log blockers in Open_Questions, then review the Before you submit checklist on Intake_Guide.",
        ],
    )


def main() -> None:
    out_dir = REPO_ROOT / "workbooks"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / OUT_NAME

    wb = Workbook()
    add_readme_sheet(wb)
    add_table_index_sheet(wb)
    add_lookup_lists_sheet(wb)
    add_intake_guide_sheet(wb)

    for sheet_name, table_name, headers, rows in INTAKE_TABLES:
        ws = wb.create_sheet(sheet_name)
        write_table_sheet(ws, headers, rows, table_name)

    wb.save(out_path)
    print(f"Wrote workbook: {out_path}")
    for sheet_name, table_name, _, _ in INTAKE_TABLES:
        print(f"  - {sheet_name} -> {table_name}")


if __name__ == "__main__":
    main()
