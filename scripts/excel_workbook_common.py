"""Shared helpers for CHI Excel workbook generators.

See docs/excel-workbook-generation-rules.md before changing Table or filter logic.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
MODEL_FILL = PatternFill("solid", fgColor="E2EFDA")
README_TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")

# Excel table naming: chi_{artifact} in snake_case, unique, no spaces.
# Matches governed ddc-* parquet artifacts where applicable.


def style_header_row(ws: Worksheet, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def autosize_columns(ws: Worksheet, max_width: int = 38, min_width: int = 12) -> None:
    for idx, col in enumerate(ws.columns, 1):
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, min_width), max_width)


def add_excel_table(ws: Worksheet, table_name: str) -> None:
    """Register the sheet data range as a named Excel Table."""
    if ws.max_row < 2 or ws.max_column < 1:
        return
    end_col = get_column_letter(ws.max_column)
    ref = f"A1:{end_col}{ws.max_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"
    # Do not set ws.auto_filter here — the Table definition already includes
    # autoFilter; a worksheet-level duplicate triggers Excel's repair dialog.


def table_lookup_formula(table_name: str, key_column: str, value_column: str) -> str:
    """INDEX/MATCH over Excel structured references (requires real Table objects)."""
    return (
        f'=IF($B$3="","",IFERROR(INDEX({table_name}[{value_column}],'
        f'MATCH($B$3,{table_name}[{key_column}],0)),""))'
    )


def add_list_validation(
    ws: Worksheet,
    column_name: str,
    list_range: str,
    headers: list[str],
    max_row: int = 500,
) -> None:
    if column_name not in headers:
        return
    col_idx = headers.index(column_name) + 1
    col_letter = get_column_letter(col_idx)
    dv = DataValidation(type="list", formula1=list_range, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max_row}")


def write_readme_sheet(
    ws: Worksheet,
    title: str,
    rows: list[tuple[str, str]],
    start_here_steps: list[str] | None = None,
) -> None:
    ws.append((title,))
    ws.append(())
    if start_here_steps:
        ws.append(("Start here", ""))
        for step_number, step in enumerate(start_here_steps, 1):
            ws.append((f"  {step_number}.", step))
        ws.append(())
    for left, right in rows:
        ws.append((left, right))
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].fill = README_TITLE_FILL
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
